"""
Flask 应用入口
提供 Web UI 和 API 接口
"""
import os
import io
import json
import logging
import threading
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file

import storage
import scheduler

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger(__name__)

app = Flask(__name__)


# ======== 页面路由 ========

@app.route("/")
def index():
    return render_template("index.html")


# ======== API 路由 ========

@app.route("/api/scrape", methods=["POST"])
def api_scrape():
    """手动触发爬取"""
    data = request.get_json() or {}
    categories = data.get("categories", ["工程", "服务"])
    company = data.get("company", "")
    sources = data.get("sources", None)

    if scheduler.scrape_status["running"]:
        return jsonify({"success": False, "message": "爬取任务正在执行中，请稍后再试"})

    def do_scrape():
        scheduler.run_scrape_job(
            categories=categories,
            company=company or None,
            sources=sources
        )

    thread = threading.Thread(target=do_scrape, daemon=True)
    thread.start()

    return jsonify({"success": True, "message": "爬取任务已启动"})


@app.route("/api/status")
def api_status():
    """获取爬取状态"""
    return jsonify(scheduler.get_status())


@app.route("/api/announcements")
def api_announcements():
    """查询公告列表"""
    page = request.args.get("page", 1, type=int)
    page_size = request.args.get("page_size", 20, type=int)
    category = request.args.get("category", "")
    company = request.args.get("company", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    keyword = request.args.get("keyword", "")
    source = request.args.get("source", "")

    items, total = storage.query_announcements(
        page=page,
        page_size=page_size,
        category=category or None,
        company=company or None,
        date_from=date_from or None,
        date_to=date_to or None,
        keyword=keyword or None,
        source=source or None,
    )

    return jsonify({
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    })


@app.route("/api/companies")
def api_companies():
    """搜索公司列表"""
    q = request.args.get("q", "")
    results = storage.search_companies(q)
    return jsonify(results)


@app.route("/api/export")
def api_export():
    """导出 Excel"""
    category = request.args.get("category", "")
    company = request.args.get("company", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    keyword = request.args.get("keyword", "")
    source = request.args.get("source", "")

    items, _ = storage.query_announcements(
        page=1,
        page_size=10000,
        category=category or None,
        company=company or None,
        date_from=date_from or None,
        date_to=date_to or None,
        keyword=keyword or None,
        source=source or None,
    )

    # 生成 Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "采购公告"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")

    headers = [
        "序号", "采购项目名称", "采购类型", "标包/标的", "预计采购金额",
        "招标人", "招标方式", "采购文件获取开始时间", "采购文件获取结束时间",
        "响应文件递交截止时间", "项目链接", "信息来源", "发布日期", "类别"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=item.get("title", ""))
        ws.cell(row=row_idx, column=3, value=item.get("announcement_type", ""))
        ws.cell(row=row_idx, column=4, value=item.get("bid_packages", ""))
        ws.cell(row=row_idx, column=5, value=item.get("estimated_amount", ""))
        ws.cell(row=row_idx, column=6, value=item.get("tenderer", "") or item.get("company", ""))
        ws.cell(row=row_idx, column=7, value=item.get("bidding_method", ""))
        ws.cell(row=row_idx, column=8, value=item.get("reg_start_time", ""))
        ws.cell(row=row_idx, column=9, value=item.get("reg_end_time", ""))
        ws.cell(row=row_idx, column=10, value=item.get("bid_deadline", ""))
        ws.cell(row=row_idx, column=11, value=item.get("url", ""))
        ws.cell(row=row_idx, column=12, value=item.get("source", ""))
        ws.cell(row=row_idx, column=13, value=item.get("publish_date", ""))
        ws.cell(row=row_idx, column=14, value=item.get("category", ""))

    # 设置列宽
    col_widths = {
        "A": 6, "B": 55, "C": 14, "D": 30, "E": 16,
        "F": 25, "G": 16, "H": 20, "I": 20,
        "J": 20, "K": 50, "L": 28, "M": 12, "N": 8
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"采购公告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/api/settings", methods=["GET", "POST"])
def api_settings():
    """读取/保存设置"""
    if request.method == "GET":
        return jsonify({
            "webhook_key": storage.get_setting("webhook_key", ""),
            "categories": storage.get_setting("categories", "工程,服务"),
            "filter_company": storage.get_setting("filter_company", ""),
            "schedule_hour": int(storage.get_setting("schedule_hour", "12")),
            "schedule_minute": int(storage.get_setting("schedule_minute", "0")),
            "max_pages": int(storage.get_setting("max_pages", "5")),
            "scrape_days": int(storage.get_setting("scrape_days", "3")),
            "scrape_sources": storage.get_setting("scrape_sources", "bidding.csg.cn,ecsg.com.cn"),
        })

    data = request.get_json() or {}

    if "webhook_key" in data:
        storage.save_setting("webhook_key", data["webhook_key"])
    if "categories" in data:
        storage.save_setting("categories", data["categories"])
    if "filter_company" in data:
        storage.save_setting("filter_company", data["filter_company"])
    if "max_pages" in data:
        storage.save_setting("max_pages", str(data["max_pages"]))
    if "scrape_days" in data:
        storage.save_setting("scrape_days", str(data["scrape_days"]))
    if "scrape_sources" in data:
        storage.save_setting("scrape_sources", data["scrape_sources"])

    if "schedule_hour" in data or "schedule_minute" in data:
        hour = int(data.get("schedule_hour", storage.get_setting("schedule_hour", "12")))
        minute = int(data.get("schedule_minute", storage.get_setting("schedule_minute", "0")))
        scheduler.update_schedule(hour, minute)

    return jsonify({"success": True, "message": "设置已保存"})


# ======== 启动 ========

if __name__ == "__main__":
    logger.info("🚀 南网采购平台爬取系统启动中...")
    scheduler.start_scheduler()
    app.run(host="0.0.0.0", port=5000, debug=False)
