"""
Excel 导出模块
- generate_excel(items) → BytesIO  供 API 下载和文件保存共用
- export_to_file(items, export_dir) → filepath  保存到指定目录
"""
import io
import os
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)


def generate_excel(items):
    """
    将公告列表生成 Excel，返回 BytesIO 对象
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "采购公告"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")

    headers = [
        "序号", "采购项目名称", "采购类型", "标包/标的", "预计采购金额",
        "招标人", "招标方式", "采购文件获取开始时间", "采购文件获取结束时间",
        "响应文件递交截止时间", "项目链接", "信息来源", "发布日期", "类别"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=item.get("title", ""))
        ws.cell(row=row_idx, column=3, value=item.get("announcement_type", ""))
        ws.cell(row=row_idx, column=4, value=item.get("bid_packages", ""))
        ws.cell(row=row_idx, column=5, value=item.get("estimated_amount", ""))
        ws.cell(row=row_idx, column=6, value=item.get("tenderer", "") or item.get("company", ""))
        ws.cell(row=row_idx, column=7, value=item.get("bidding_method", ""))
        ws.cell(row=row_idx, column=8, value=item.get("reg_start_time", ""))
        ws.cell(row=row_idx, column=9, value=item.get("reg_end_time", ""))
        ws.cell(row=row_idx, column=10, value=item.get("bid_deadline", ""))
        ws.cell(row=row_idx, column=11, value=item.get("url", ""))
        ws.cell(row=row_idx, column=12, value=item.get("source", ""))
        ws.cell(row=row_idx, column=13, value=item.get("publish_date", ""))
        ws.cell(row=row_idx, column=14, value=item.get("category", ""))

    # 设置列宽
    col_widths = {
        "A": 6, "B": 55, "C": 14, "D": 30, "E": 16,
        "F": 25, "G": 16, "H": 20, "I": 20,
        "J": 20, "K": 50, "L": 28, "M": 12, "N": 8
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_file(items, export_dir):
    """
    将公告列表导出为 Excel 文件并保存到指定目录
    返回保存的文件路径
    """
    if not items:
        logger.info("没有数据需要导出")
        return None

    # 确保目录存在
    os.makedirs(export_dir, exist_ok=True)

    filename = f"采购公告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(export_dir, filename)

    excel_data = generate_excel(items)

    with open(filepath, "wb") as f:
        f.write(excel_data.getvalue())

    logger.info(f"Excel 已导出: {filepath}")
    return filepath
